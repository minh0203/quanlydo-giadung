from PyQt5.QtWidgets import (QMessageBox, QTableWidgetItem, QLineEdit, QPushButton, 
                           QComboBox, QTextEdit, QRadioButton, QDateEdit, QFileDialog)
from PyQt5.QtCore import Qt, QDate
from datetime import datetime
from models.employee import Employee
from models.database import Database


class EmployeeValidator:
    """Lớp xác thực dữ liệu nhân viên"""

    @staticmethod
    def validate_full_name(full_name):
        """Kiểm tra tên đầy đủ"""
        return bool(full_name and len(full_name.strip()) >= 3)

    @staticmethod
    def validate_email(email):
        """Kiểm tra email"""
        if not email:
            return True
        import re
        pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        return bool(re.match(pattern, email))

    @staticmethod
    def validate_phone(phone):
        """Kiểm tra số điện thoại"""
        if not phone:
            return True
        import re
        pattern = r'^\+?[0-9]{9,15}$'
        return bool(re.match(pattern, phone.replace(" ", "").replace("-", "")))

    @staticmethod
    def validate_identity_card(identity_card):
        """Kiểm tra CMND/CCCD"""
        if not identity_card:
            return True
        return len(identity_card.strip()) in [9, 12]

    @staticmethod
    def validate_base_salary(base_salary):
        """Kiểm tra lương cơ bản"""
        return base_salary >= 0

    @staticmethod
    def validate_commission_rate(commission_rate):
        """Kiểm tra tỷ lệ hoa hồng"""
        return 0 <= commission_rate <= 100

    @staticmethod
    def validate_username(username):
        """Kiểm tra tài khoản"""
        return bool(username and len(username.strip()) >= 3)

    @staticmethod
    def validate_password(password):
        """Kiểm tra mật khẩu"""
        return bool(password and len(password) >= 6)


class EmployeeController:
    """Controller xử lý logic quản lý nhân viên"""
    
    def __init__(self, view):
        self.view = view
        self.current_employee = None
        self.initialize_ui()
        self.setup_connections()
        self.load_employees()
    
    def initialize_ui(self):
        """Khởi tạo các giá trị UI"""
        # Populate combo boxes
        if hasattr(self.view, "cboRole"):
            self.view.cboRole.clear()
            self.view.cboRole.addItems([
                "Admin (Quản trị hệ thống)",
                "Sale (Nhân viên bán hàng)",
                "Warehouse (Quản lý kho)",
                "Accountant (Kế toán)"
            ])
        
        if hasattr(self.view, "cboStatus"):
            self.view.cboStatus.clear()
            self.view.cboStatus.addItems([
                "✅ Đang làm",
                "⏸️ Tạm nghỉ",
                "❌ Đã nghỉ"
            ])
        
        if hasattr(self.view, "cboRoleFilter"):
            self.view.cboRoleFilter.clear()
            self.view.cboRoleFilter.addItems([
                "Tất cả",
                "Admin",
                "Sale",
                "Warehouse",
                "Accountant"
            ])
        
        if hasattr(self.view, "cboStatusFilter"):
            self.view.cboStatusFilter.clear()
            self.view.cboStatusFilter.addItems([
                "Tất cả",
                "Đang làm",
                "Tạm nghỉ",
                "Đã nghỉ"
            ])

    def setup_connections(self):
        """Thiết lập kết nối các signal"""
        # Kết nối nút thêm nhân viên
        if hasattr(self.view, "btnAdd"):
            self.view.btnAdd.clicked.connect(self.show_add_form)

        # Kết nối nút lưu nhân viên
        if hasattr(self.view, "btnSave"):
            self.view.btnSave.clicked.connect(self.save_employee)

        # Kết nối nút cập nhật nhân viên
        if hasattr(self.view, "btnUpdate"):
            self.view.btnUpdate.clicked.connect(self.update_employee)

        # Kết nối nút xóa nhân viên
        if hasattr(self.view, "btnDelete"):
            self.view.btnDelete.clicked.connect(self.delete_employee)

        # Kết nối nút xóa form
        if hasattr(self.view, "btnClear"):
            self.view.btnClear.clicked.connect(self.clear_form)

        # Kết nối nút tìm kiếm
        if hasattr(self.view, "btnSearch"):
            self.view.btnSearch.clicked.connect(self.search_employees)

        # Kết nối nút xuất Excel
        if hasattr(self.view, "btnExportExcel"):
            self.view.btnExportExcel.clicked.connect(self.export_excel)

        # Kết nối ô tìm kiếm
        if hasattr(self.view, "txtSearch"):
            self.view.txtSearch.textChanged.connect(self.on_search_text_changed)

        # Kết nối combo lọc theo vai trò
        if hasattr(self.view, "cboRoleFilter"):
            self.view.cboRoleFilter.currentTextChanged.connect(self.apply_filters)

        # Kết nối combo lọc theo trạng thái
        if hasattr(self.view, "cboStatusFilter"):
            self.view.cboStatusFilter.currentTextChanged.connect(self.apply_filters)

        # Kết nối bảng nhân viên
        if hasattr(self.view, "tableEmployees"):
            self.view.tableEmployees.itemSelectionChanged.connect(self.on_employee_selected)

        # Kết nối nút chọn ảnh
        if hasattr(self.view, "btnSelectAvatar"):
            self.view.btnSelectAvatar.clicked.connect(self.select_avatar)

    def load_employees(self):
        """Tải danh sách nhân viên"""
        try:
            employees = Employee.get_all()
            self.display_employees(employees)
        except Exception as e:
            QMessageBox.warning(None, "Lỗi", f"Không thể tải danh sách nhân viên: {str(e)}")

    def display_employees(self, employees):
        """Hiển thị danh sách nhân viên lên bảng"""
        if not hasattr(self.view, "tableEmployees"):
            return

        table = self.view.tableEmployees
        table.setRowCount(len(employees))
        table.setColumnCount(10)
        table.setHorizontalHeaderLabels([
            "ID", "Mã NV", "Họ tên", "Giới tính", "Điện thoại",
            "Email", "Vai trò", "Trạng thái", "Lương CB", "Ngày vào"
        ])

        for row, employee in enumerate(employees):
            table.setItem(row, 0, QTableWidgetItem(employee.employee_id))
            table.setItem(row, 1, QTableWidgetItem(employee.employee_id))
            table.setItem(row, 2, QTableWidgetItem(employee.full_name))
            table.setItem(row, 3, QTableWidgetItem(employee.gender))
            table.setItem(row, 4, QTableWidgetItem(employee.phone))
            table.setItem(row, 5, QTableWidgetItem(employee.email))
            table.setItem(row, 6, QTableWidgetItem(employee.role))
            table.setItem(row, 7, QTableWidgetItem(employee.status))
            table.setItem(row, 8, QTableWidgetItem(f"{employee.base_salary:,.0f} VNĐ"))
            table.setItem(row, 9, QTableWidgetItem(employee.hire_date))

        # Điều chỉnh kích thước cột
        table.resizeColumnsToContents()

    def show_add_form(self):
        """Hiển thị form thêm nhân viên mới"""
        self.clear_form()
        self.current_employee = None
        # Disable mã nhân viên vì tự động tạo
        self.set_field_enabled("txtEmployeeCode", False)
        if hasattr(self.view, "btnSave"):
            self.view.btnSave.setEnabled(True)
        if hasattr(self.view, "btnUpdate"):
            self.view.btnUpdate.setEnabled(False)

    def save_employee(self):
        """Lưu nhân viên (thêm mới hoặc cập nhật)"""
        try:
            # Lấy dữ liệu từ form
            full_name = self.get_text_field("txtFullName")
            gender = self.get_gender()
            birth_date = self.get_date_field("dateBirthDate")
            phone = self.get_text_field("txtPhone")
            email = self.get_text_field("txtEmail")
            identity_card = self.get_text_field("txtIdentityCard")
            address = self.get_text_field("txtAddress")
            hire_date = self.get_date_field("dateHireDate")
            role = self.get_combo_field("cboRole")
            base_salary = self.get_float_field("txtBaseSalary")
            commission_rate = self.get_float_field("txtCommissionRate")
            username = self.get_text_field("txtUsername")
            password = self.get_text_field("txtPassword")
            status = self.get_combo_field("cboStatus")
            note = self.get_text_field("txtNote")

            # Kiểm tra dữ liệu bắt buộc
            if not EmployeeValidator.validate_full_name(full_name):
                QMessageBox.warning(None, "Cảnh báo", "Vui lòng nhập tên nhân viên (tối thiểu 3 ký tự)!")
                return

            if not EmployeeValidator.validate_email(email):
                QMessageBox.warning(None, "Cảnh báo", "Email không hợp lệ!")
                return

            if not EmployeeValidator.validate_phone(phone):
                QMessageBox.warning(None, "Cảnh báo", "Số điện thoại không hợp lệ!")
                return

            if not EmployeeValidator.validate_identity_card(identity_card):
                QMessageBox.warning(None, "Cảnh báo", "CMND/CCCD phải có 9 hoặc 12 ký tự!")
                return

            if not EmployeeValidator.validate_base_salary(base_salary):
                QMessageBox.warning(None, "Cảnh báo", "Lương cơ bản không thể âm!")
                return

            if not EmployeeValidator.validate_commission_rate(commission_rate):
                QMessageBox.warning(None, "Cảnh báo", "Tỷ lệ hoa hồng phải từ 0 đến 100%!")
                return

            if self.current_employee:
                # Cập nhật nhân viên hiện tại
                self.current_employee.full_name = full_name
                self.current_employee.gender = gender
                self.current_employee.birth_date = birth_date
                self.current_employee.phone = phone
                self.current_employee.email = email
                self.current_employee.identity_card = identity_card
                self.current_employee.address = address
                self.current_employee.hire_date = hire_date
                self.current_employee.role = role
                self.current_employee.base_salary = base_salary
                self.current_employee.commission_rate = commission_rate
                self.current_employee.username = username
                self.current_employee.password = password
                self.current_employee.status = status
                self.current_employee.note = note

                self.current_employee.update()
                QMessageBox.information(None, "Thành công", "Đã cập nhật nhân viên thành công!")
            else:
                # Kiểm tra trùng lặp
                if email and Employee.check_duplicate_email(email):
                    QMessageBox.warning(None, "Cảnh báo", "Email này đã tồn tại trong hệ thống!")
                    return

                if identity_card and Employee.check_duplicate_identity_card(identity_card):
                    QMessageBox.warning(None, "Cảnh báo", "CMND/CCCD này đã tồn tại trong hệ thống!")
                    return

                if username and Employee.check_duplicate_username(username):
                    QMessageBox.warning(None, "Cảnh báo", "Tài khoản này đã tồn tại trong hệ thống!")
                    return

                # Tạo nhân viên mới
                employee = Employee.create(
                    full_name=full_name,
                    gender=gender,
                    birth_date=birth_date,
                    phone=phone,
                    email=email,
                    identity_card=identity_card,
                    address=address,
                    hire_date=hire_date,
                    role=role,
                    base_salary=base_salary,
                    commission_rate=commission_rate,
                    username=username,
                    password=password,
                    status=status,
                    note=note
                )
                QMessageBox.information(None, "Thành công", f"Đã thêm nhân viên thành công! Mã nhân viên: {employee.employee_id}")

            self.clear_form()
            self.load_employees()

        except Exception as e:
            QMessageBox.warning(None, "Lỗi", f"Không thể lưu nhân viên: {str(e)}")

    def update_employee(self):
        """Cập nhật nhân viên được chọn"""
        if not self.current_employee:
            QMessageBox.warning(None, "Cảnh báo", "Vui lòng chọn nhân viên cần cập nhật!")
            return

        # Enable các field để chỉnh sửa
        if hasattr(self.view, "btnSave"):
            self.view.btnSave.setEnabled(True)
        if hasattr(self.view, "btnUpdate"):
            self.view.btnUpdate.setEnabled(False)

    def delete_employee(self):
        """Xóa nhân viên"""
        if not self.current_employee:
            QMessageBox.warning(None, "Cảnh báo", "Vui lòng chọn nhân viên cần xóa!")
            return

        reply = QMessageBox.question(
            None, "Xác nhận",
            f"Bạn có chắc muốn xóa nhân viên \'{self.current_employee.full_name}\'?",
            QMessageBox.Yes | QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            try:
                self.current_employee.delete()
                QMessageBox.information(None, "Thành công", "Đã xóa nhân viên thành công!")
                self.clear_form()
                self.load_employees()
            except Exception as e:
                QMessageBox.warning(None, "Lỗi", f"Không thể xóa nhân viên: {str(e)}")

    def search_employees(self):
        """Tìm kiếm nhân viên"""
        keyword = self.get_text_field("txtSearch").strip()
        if keyword:
            try:
                employees = Employee.search(keyword)
                self.display_employees(employees)
            except Exception as e:
                QMessageBox.warning(None, "Lỗi", f"Không thể tìm kiếm: {str(e)}")
        else:
            self.load_employees()

    def on_search_text_changed(self):
        """Xử lý khi text tìm kiếm thay đổi"""
        if hasattr(self.view, "txtSearch"):
            if not self.view.txtSearch.text().strip():
                self.load_employees()

    def apply_filters(self):
        """Áp dụng bộ lọc theo vai trò và trạng thái"""
        try:
            role_filter = self.get_combo_field("cboRoleFilter")
            status_filter = self.get_combo_field("cboStatusFilter")

            employees = Employee.get_all()

            if role_filter:
                employees = [e for e in employees if e.role == role_filter]

            if status_filter:
                employees = [e for e in employees if e.status == status_filter]

            self.display_employees(employees)
        except Exception as e:
            QMessageBox.warning(None, "Lỗi", f"Không thể áp dụng bộ lọc: {str(e)}")

    def on_employee_selected(self):
        """Xử lý khi chọn nhân viên từ bảng"""
        if not hasattr(self.view, "tableEmployees"):
            return

        current_row = self.view.tableEmployees.currentRow()
        if current_row >= 0:
            employee_id = self.view.tableEmployees.item(current_row, 0).text()
            self.current_employee = Employee.get_by_id(employee_id)
            if self.current_employee:
                self.fill_form()
                self.update_statistics()

    def fill_form(self):
        """Điền dữ liệu nhân viên vào form"""
        if not self.current_employee:
            return

        self.set_text_field("txtEmployeeCode", self.current_employee.employee_id)
        self.set_text_field("txtFullName", self.current_employee.full_name)
        self.set_gender(self.current_employee.gender)
        self.set_date_field("dateBirthDate", self.current_employee.birth_date)
        self.set_text_field("txtPhone", self.current_employee.phone)
        self.set_text_field("txtEmail", self.current_employee.email)
        self.set_text_field("txtIdentityCard", self.current_employee.identity_card)
        self.set_text_field("txtAddress", self.current_employee.address)
        self.set_date_field("dateHireDate", self.current_employee.hire_date)
        self.set_combo_field("cboRole", self.current_employee.role)
        self.set_text_field("txtBaseSalary", str(int(self.current_employee.base_salary)))
        self.set_text_field("txtCommissionRate", str(self.current_employee.commission_rate))
        self.set_text_field("txtUsername", self.current_employee.username)
        self.set_text_field("txtPassword", self.current_employee.password)
        self.set_combo_field("cboStatus", self.current_employee.status)
        self.set_text_field("txtNote", self.current_employee.note)

        # Disable các nút
        if hasattr(self.view, "btnSave"):
            self.view.btnSave.setEnabled(False)
        if hasattr(self.view, "btnUpdate"):
            self.view.btnUpdate.setEnabled(True)
    
    def update_statistics(self):
        """Cập nhật thống kê nhân viên"""
        if not self.current_employee:
            return
        
        try:
            age = self.current_employee.get_age()
            working_days = self.current_employee.get_working_days()
            working_years = working_days // 365
            working_months = (working_days % 365) // 30
            
            stats_text = f"""
            📊 THỐNG KÊ NHÂN VIÊN
            {'='*50}
            
            👤 Thông tin cơ bản:
            • Mã NV: {self.current_employee.employee_id}
            • Tên: {self.current_employee.full_name}
            • Tuổi: {age if age else 'N/A'} tuổi
            • Giới tính: {self.current_employee.gender}
            
            💼 Thông tin làm việc:
            • Chức vụ: {self.current_employee.role}
            • Ngày vào: {self.current_employee.hire_date}
            • Thời gian làm việc: {working_years} năm {working_months} tháng
            • Trạng thái: {self.current_employee.status}
            
            💰 Thông tin lương:
            • Lương cơ bản: {self.current_employee.base_salary:,.0f} VNĐ
            • Hoa hồng: {self.current_employee.commission_rate}%
            • Lương theo giờ: {self.current_employee.base_salary / 160:,.0f} VNĐ
            
            📞 Liên hệ:
            • Email: {self.current_employee.email}
            • Điện thoại: {self.current_employee.phone}
            • Địa chỉ: {self.current_employee.address}
            """
            
            if hasattr(self.view, "lblStatsInfo"):
                self.view.lblStatsInfo.setText(stats_text)
        except Exception as e:
            if hasattr(self.view, "lblStatsInfo"):
                self.view.lblStatsInfo.setText(f"Lỗi: {str(e)}")

    def clear_form(self):
        """Xóa dữ liệu trong form"""
        self.set_text_field("txtEmployeeCode", "")
        self.set_text_field("txtFullName", "")
        self.set_gender("Nam")
        self.set_date_field("dateBirthDate", "1990-01-01")
        self.set_text_field("txtPhone", "")
        self.set_text_field("txtEmail", "")
        self.set_text_field("txtIdentityCard", "")
        self.set_text_field("txtAddress", "")
        self.set_date_field("dateHireDate", datetime.now().strftime("%Y-%m-%d"))
        self.set_combo_field("cboRole", "Sale")
        self.set_text_field("txtBaseSalary", "0.0")
        self.set_text_field("txtCommissionRate", "0.0")
        self.set_text_field("txtUsername", "")
        self.set_text_field("txtPassword", "")
        self.set_combo_field("cboStatus", "Đang làm")
        self.set_text_field("txtNote", "")
        self.set_text_field("lblAvatar", "👤")

        # Enable các field
        self.set_field_enabled("txtEmployeeCode", True)

        # Reset trạng thái nút
        if hasattr(self.view, "btnSave"):
            self.view.btnSave.setEnabled(True)
        if hasattr(self.view, "btnUpdate"):
            self.view.btnUpdate.setEnabled(False)

        self.current_employee = None

    def select_avatar(self):
        """Chọn ảnh đại diện"""
        file_dialog = QFileDialog()
        image_file, _ = file_dialog.getOpenFileName(
            None, "Chọn ảnh đại diện", "",
            "Tệp ảnh (*.png *.jpg *.jpeg *.bmp);;Tất cả tệp (*)"
        )
        if image_file:
            self.set_text_field("lblAvatar", "📸")
            # Có thể lưu đường dẫn ảnh vào một field ẩn hoặc xử lý khác

    def export_excel(self):
        """Xuất dữ liệu nhân viên ra Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            employees = Employee.get_all()
            if not employees:
                QMessageBox.warning(None, "Cảnh báo", "Không có nhân viên nào để xuất!")
                return

            # Tạo workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Nhân viên"

            # Tạo header
            headers = ["Mã NV", "Họ tên", "Giới tính", "Ngày sinh", "Điện thoại", "Email",
                      "CMND/CCCD", "Địa chỉ", "Ngày vào", "Vai trò", "Lương CB", "Hoa hồng", "Trạng thái"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")

            # Điền dữ liệu
            for row_idx, employee in enumerate(employees, 2):
                ws.cell(row=row_idx, column=1).value = employee.employee_id
                ws.cell(row=row_idx, column=2).value = employee.full_name
                ws.cell(row=row_idx, column=3).value = employee.gender
                ws.cell(row=row_idx, column=4).value = employee.birth_date
                ws.cell(row=row_idx, column=5).value = employee.phone
                ws.cell(row=row_idx, column=6).value = employee.email
                ws.cell(row=row_idx, column=7).value = employee.identity_card
                ws.cell(row=row_idx, column=8).value = employee.address
                ws.cell(row=row_idx, column=9).value = employee.hire_date
                ws.cell(row=row_idx, column=10).value = employee.role
                ws.cell(row=row_idx, column=11).value = employee.base_salary
                ws.cell(row=row_idx, column=12).value = employee.commission_rate
                ws.cell(row=row_idx, column=13).value = employee.status

            # Điều chỉnh độ rộng cột
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Lưu file
            from datetime import datetime
            filename = f"nhan_vien_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(filename)
            QMessageBox.information(None, "Thành công", f"Đã xuất file: {filename}")

        except ImportError:
            QMessageBox.warning(None, "Lỗi", "Vui lòng cài đặt thư viện openpyxl: pip install openpyxl")
        except Exception as e:
            QMessageBox.warning(None, "Lỗi", f"Không thể xuất file: {str(e)}")

    # Helper methods
    def get_text_field(self, field_name):
        """Lấy text từ field"""
        if hasattr(self.view, field_name):
            field = getattr(self.view, field_name)
            if isinstance(field, QLineEdit):
                return field.text().strip()
            elif isinstance(field, QTextEdit):
                return field.toPlainText().strip()
            elif isinstance(field, QComboBox):
                return field.currentText().strip()
        return ""

    def set_text_field(self, field_name, value):
        """Đặt text cho field"""
        if hasattr(self.view, field_name):
            field = getattr(self.view, field_name)
            if isinstance(field, QLineEdit):
                field.setText(str(value))
            elif isinstance(field, QTextEdit):
                field.setPlainText(str(value))
            elif isinstance(field, QComboBox):
                field.setCurrentText(str(value))

    def get_combo_field(self, field_name):
        """Lấy giá trị từ combo box"""
        if hasattr(self.view, field_name):
            field = getattr(self.view, field_name)
            if isinstance(field, QComboBox):
                text = field.currentText().strip()
                
                # Mapping ngược cho các combo box đặc biệt
                role_mapping = {
                    "Admin (Quản trị hệ thống)": "Admin",
                    "Sale (Nhân viên bán hàng)": "Sale",
                    "Warehouse (Quản lý kho)": "Warehouse",
                    "Accountant (Kế toán)": "Accountant"
                }
                
                status_mapping = {
                    "✅ Đang làm": "Đang làm",
                    "⏸️ Tạm nghỉ": "Tạm nghỉ",
                    "❌ Đã nghỉ": "Đã nghỉ"
                }
                
                # Chuyển đổi giá trị nếu cần
                if field_name == "cboRole" and text in role_mapping:
                    return role_mapping[text]
                elif field_name == "cboStatus" and text in status_mapping:
                    return status_mapping[text]
                
                return text
        return ""

    def set_combo_field(self, field_name, value):
        """Đặt giá trị cho combo box"""
        if hasattr(self.view, field_name):
            field = getattr(self.view, field_name)
            if isinstance(field, QComboBox):
                # Mapping cho các combo box đặc biệt
                role_mapping = {
                    "Admin": "Admin (Quản trị hệ thống)",
                    "Sale": "Sale (Nhân viên bán hàng)",
                    "Warehouse": "Warehouse (Quản lý kho)",
                    "Accountant": "Accountant (Kế toán)"
                }
                
                status_mapping = {
                    "Đang làm": "✅ Đang làm",
                    "Tạm nghỉ": "⏸️ Tạm nghỉ",
                    "Đã nghỉ": "❌ Đã nghỉ"
                }
                
                # Chuyển đổi giá trị nếu cần
                display_value = str(value)
                if field_name == "cboRole" and str(value) in role_mapping:
                    display_value = role_mapping[str(value)]
                elif field_name == "cboStatus" and str(value) in status_mapping:
                    display_value = status_mapping[str(value)]
                
                # Tìm và set
                index = field.findText(display_value)
                if index >= 0:
                    field.setCurrentIndex(index)
                else:
                    field.setCurrentText(display_value)

    def get_date_field(self, field_name):
        """Lấy ngày từ date field"""
        if hasattr(self.view, field_name):
            field = getattr(self.view, field_name)
            if isinstance(field, QDateEdit):
                return field.date().toString("yyyy-MM-dd")
        return ""

    def set_date_field(self, field_name, date_string):
        """Đặt ngày cho date field"""
        if hasattr(self.view, field_name):
            field = getattr(self.view, field_name)
            if isinstance(field, QDateEdit):
                try:
                    if date_string:
                        date_obj = QDate.fromString(date_string, "yyyy-MM-dd")
                        if date_obj.isValid():
                            field.setDate(date_obj)
                    else:
                        field.setDate(QDate(1990, 1, 1))
                except:
                    field.setDate(QDate(1990, 1, 1))

    def get_gender(self):
        """Lấy giới tính từ radio button"""
        if hasattr(self.view, "rdMale") and self.view.rdMale.isChecked():
            return "Nam"
        elif hasattr(self.view, "rdFemale") and self.view.rdFemale.isChecked():
            return "Nữ"
        elif hasattr(self.view, "rdOther") and self.view.rdOther.isChecked():
            return "Khác"
        return "Nam"

    def set_gender(self, gender):
        """Đặt giới tính bằng radio button"""
        if gender == "Nữ" and hasattr(self.view, "rdFemale"):
            self.view.rdFemale.setChecked(True)
        elif gender == "Khác" and hasattr(self.view, "rdOther"):
            self.view.rdOther.setChecked(True)
        elif hasattr(self.view, "rdMale"):
            self.view.rdMale.setChecked(True)

    def set_field_enabled(self, field_name, enabled):
        """Enable/disable field"""
        if hasattr(self.view, field_name):
            field = getattr(self.view, field_name)
            field.setEnabled(enabled)

    def get_int_field(self, field_name):
        """Lấy giá trị int từ field"""
        text = self.get_text_field(field_name)
        try:
            return int(text) if text else 0
        except ValueError:
            return 0

    def get_float_field(self, field_name):
        """Lấy giá trị float từ field"""
        text = self.get_text_field(field_name)
        try:
            return float(text) if text else 0.0
        except ValueError:
            return 0.0