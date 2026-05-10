# -*- coding: utf-8 -*-
from PyQt5.QtWidgets import QMessageBox, QTableWidgetItem, QFileDialog
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QColor, QFont
from models.database import Database
from models.product import Product
from models.product_category import ProductCategory
import csv
from datetime import datetime
from pathlib import Path


class InventoryReportController:
    """Controller xử lý logic báo cáo tồn kho"""

    def __init__(self, view):
        self.view = view
        self.current_data = []  # Lưu trữ dữ liệu hiện tại để sử dụng cho các chức năng khác
        try:
            self.setup_connections()
            self.load_categories_and_brands()
            self.load_inventory_data()
        except Exception as e:
            print(f"Lỗi khi khởi tạo InventoryReportController: {str(e)}")

    def setup_connections(self):
        """Thiết lập kết nối các signal"""
        # Kết nối nút tìm kiếm/xem báo cáo
        if hasattr(self.view, 'btnSearch'):
            self.view.btnSearch.clicked.connect(self.search_inventory)

        # Kết nối nút xuất báo cáo Excel
        if hasattr(self.view, 'btnExport'):
            self.view.btnExport.clicked.connect(self.export_to_excel)

        # Kết nối nút in báo cáo
        if hasattr(self.view, 'btnPrint'):
            self.view.btnPrint.clicked.connect(self.print_report)

        # Kết nối nút đặt lại filter
        if hasattr(self.view, 'btnReset'):
            self.view.btnReset.clicked.connect(self.reset_filters)

    def load_categories_and_brands(self):
        """Tải danh sách danh mục và thương hiệu từ database"""
        try:
            # Tải danh mục
            if hasattr(self.view, 'cboCategory'):
                categories = ProductCategory.get_all()
                self.view.cboCategory.clear()
                self.view.cboCategory.addItem("Tất cả danh mục")
                for category in categories:
                    if category:  # Bỏ qua các danh mục trống
                        self.view.cboCategory.addItem(category)

            # Tải thương hiệu
            if hasattr(self.view, 'cboBrand'):
                brands = self.get_all_brands()
                self.view.cboBrand.clear()
                self.view.cboBrand.addItem("Tất cả")
                for brand in brands:
                    if brand:  # Bỏ qua các thương hiệu trống
                        self.view.cboBrand.addItem(brand)

            # Tải trạng thái tồn kho
            if hasattr(self.view, 'cboStockStatus'):
                self.view.cboStockStatus.clear()
                self.view.cboStockStatus.addItem("Tất cả")
                self.view.cboStockStatus.addItem("Còn hàng (>0)")
                self.view.cboStockStatus.addItem("Sắp hết (≤10)")
                self.view.cboStockStatus.addItem("Hết hàng (=0)")
                self.view.cboStockStatus.addItem("Tồn nhiều (>50)")

        except Exception as e:
            print(f"Lỗi khi tải danh mục và thương hiệu: {str(e)}")

    def get_all_brands(self):
        """Lấy tất cả thương hiệu từ database"""
        try:
            rows = Database.execute(
                "SELECT DISTINCT brand FROM products WHERE brand IS NOT NULL AND brand <> '' ORDER BY brand",
                fetch_all=True,
            )
            return [row[0] for row in rows] if rows else []
        except Exception as e:
            print(f"Lỗi khi lấy thương hiệu: {str(e)}")
            return []

    def load_inventory_data(self, category=None, brand=None, stock_status=None):
        """
        Tải dữ liệu tồn kho từ database
        
        Args:
            category: Danh mục lọc (None = tất cả)
            brand: Thương hiệu lọc (None = tất cả)
            stock_status: Trạng thái tồn (None = tất cả, ">0" = còn hàng, "<=10" = sắp hết, "=0" = hết hàng, ">50" = tồn nhiều)
        """
        try:
            # Lấy tất cả sản phẩm
            products = Product.get_all() or []
            
            # Lọc theo điều kiện
            filtered_products = []
            for product in products:
                if not product:
                    continue
                    
                # Lọc theo danh mục
                if category and category != "Tất cả danh mục":
                    if (product.category or "") != category:
                        continue
                
                # Lọc theo thương hiệu
                if brand and brand != "Tất cả":
                    if (product.brand or "") != brand:
                        continue
                
                # Lọc theo trạng thái tồn kho
                if stock_status and stock_status != "Tất cả":
                    quantity = int(product.quantity or 0)
                    if stock_status == "Còn hàng (>0)" and quantity <= 0:
                        continue
                    elif stock_status == "Sắp hết (≤10)" and quantity > 10:
                        continue
                    elif stock_status == "Hết hàng (=0)" and quantity != 0:
                        continue
                    elif stock_status == "Tồn nhiều (>50)" and quantity <= 50:
                        continue
                
                filtered_products.append(product)
            
            self.current_data = filtered_products
            self.display_inventory_data(filtered_products)
            self.update_statistics()
            self.display_top_inventory()

        except Exception as e:
            print(f"Lỗi khi tải dữ liệu tồn kho: {str(e)}")
            QMessageBox.warning(self.view.parent_widget, "Lỗi", f"Không thể tải dữ liệu tồn kho: {str(e)}")

    def display_inventory_data(self, data):
        """Hiển thị dữ liệu tồn kho lên bảng chi tiết"""
        if not hasattr(self.view, 'tableInventory'):
            return

        table = self.view.tableInventory
        table.setRowCount(len(data))

        # Tính tổng giá trị tồn kho trước (ngoài loop)
        total_inventory_value = sum(p.selling_price * p.quantity for p in data) if data else 1

        for row, product in enumerate(data):
            try:
                # Mã sản phẩm
                item_code = QTableWidgetItem(str(product.product_id or ""))
                item_code.setFlags(item_code.flags() & ~Qt.ItemIsEditable)
                table.setItem(row, 0, item_code)

                # Tên sản phẩm
                item_name = QTableWidgetItem(str(product.name or ""))
                item_name.setFlags(item_name.flags() & ~Qt.ItemIsEditable)
                table.setItem(row, 1, item_name)

                # Danh mục
                item_category = QTableWidgetItem(str(product.category or ""))
                item_category.setFlags(item_category.flags() & ~Qt.ItemIsEditable)
                table.setItem(row, 2, item_category)

                # Thương hiệu
                item_brand = QTableWidgetItem(str(product.brand or ""))
                item_brand.setFlags(item_brand.flags() & ~Qt.ItemIsEditable)
                table.setItem(row, 3, item_brand)

                # Giá nhập
                purchase_price = float(product.purchase_price or 0)
                item_purchase = QTableWidgetItem(f"{purchase_price:,.0f} VNĐ")
                item_purchase.setFlags(item_purchase.flags() & ~Qt.ItemIsEditable)
                item_purchase.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                table.setItem(row, 4, item_purchase)

                # Giá bán
                selling_price = float(product.selling_price or 0)
                item_selling = QTableWidgetItem(f"{selling_price:,.0f} VNĐ")
                item_selling.setFlags(item_selling.flags() & ~Qt.ItemIsEditable)
                item_selling.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                table.setItem(row, 5, item_selling)

                # Tồn kho
                quantity = int(product.quantity or 0)
                item_quantity = QTableWidgetItem(str(quantity))
                item_quantity.setFlags(item_quantity.flags() & ~Qt.ItemIsEditable)
                item_quantity.setTextAlignment(Qt.AlignCenter)
                
                # Tô màu theo tình trạng tồn kho
                if quantity == 0:
                    item_quantity.setBackground(QColor(255, 100, 100))  # Đỏ - hết hàng
                elif quantity <= 10:
                    item_quantity.setBackground(QColor(255, 200, 100))  # Cam - sắp hết
                elif quantity > 50:
                    item_quantity.setBackground(QColor(100, 200, 100))  # Xanh - tồn nhiều
                
                table.setItem(row, 6, item_quantity)

                # Giá trị tồn kho
                inventory_value = selling_price * quantity
                item_value = QTableWidgetItem(f"{inventory_value:,.0f} VNĐ")
                item_value.setFlags(item_value.flags() & ~Qt.ItemIsEditable)
                item_value.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                table.setItem(row, 7, item_value)

                # Tỷ lệ %
                percentage = (inventory_value / total_inventory_value * 100) if total_inventory_value > 0 else 0
                item_percentage = QTableWidgetItem(f"{percentage:.2f}%")
                item_percentage.setFlags(item_percentage.flags() & ~Qt.ItemIsEditable)
                item_percentage.setTextAlignment(Qt.AlignCenter)
                table.setItem(row, 8, item_percentage)

            except Exception as e:
                print(f"Lỗi khi hiển thị hàng {row}: {str(e)}")
                continue

        # Điều chỉnh kích thước cột
        table.resizeColumnsToContents()

    def update_statistics(self):
        """Cập nhật thống kê nhanh"""
        try:
            if not self.current_data:
                total_products = 0
                total_quantity = 0
                total_value = 0
            else:
                total_products = len(self.current_data)
                total_quantity = sum(int(p.quantity or 0) for p in self.current_data)
                total_value = sum(float(p.selling_price or 0) * int(p.quantity or 0) for p in self.current_data)

            # Cập nhật tổng số sản phẩm
            if hasattr(self.view, 'lblTotalProducts'):
                self.view.lblTotalProducts.setText(f"Tổng sản phẩm: {total_products}")

            # Cập nhật tổng tồn kho
            if hasattr(self.view, 'lblTotalQuantity'):
                self.view.lblTotalQuantity.setText(f"Tổng tồn kho: {total_quantity}")

            # Cập nhật tổng giá trị tồn
            if hasattr(self.view, 'lblTotalValue'):
                self.view.lblTotalValue.setText(f"Tổng giá trị: {total_value:,.0f} VNĐ")
                
        except Exception as e:
            print(f"Lỗi khi cập nhật thống kê: {str(e)}")

    def display_top_inventory(self):
        """Hiển thị top sản phẩm tồn kho nhiều nhất"""
        if not hasattr(self.view, 'tableTopInventory'):
            return

        table = self.view.tableTopInventory

        try:
            # Sắp xếp theo số lượng tồn kho giảm dần và lấy top 10
            top_products = sorted(
                [p for p in self.current_data if p], 
                key=lambda p: int(p.quantity or 0), 
                reverse=True
            )[:10]

            table.setRowCount(len(top_products))

            for row, product in enumerate(top_products):
                try:
                    # STT
                    item_stt = QTableWidgetItem(str(row + 1))
                    item_stt.setFlags(item_stt.flags() & ~Qt.ItemIsEditable)
                    item_stt.setTextAlignment(Qt.AlignCenter)
                    table.setItem(row, 0, item_stt)

                    # Tên sản phẩm
                    product_name = str(product.name or "")
                    product_brand = str(product.brand or "N/A")
                    item_name = QTableWidgetItem(f"{product_name} ({product_brand})")
                    item_name.setFlags(item_name.flags() & ~Qt.ItemIsEditable)
                    table.setItem(row, 1, item_name)

                    # Số lượng tồn
                    quantity = int(product.quantity or 0)
                    item_quantity = QTableWidgetItem(str(quantity))
                    item_quantity.setFlags(item_quantity.flags() & ~Qt.ItemIsEditable)
                    item_quantity.setTextAlignment(Qt.AlignCenter)
                    table.setItem(row, 2, item_quantity)

                    # Giá trị tồn
                    selling_price = float(product.selling_price or 0)
                    inventory_value = selling_price * quantity
                    item_value = QTableWidgetItem(f"{inventory_value:,.0f} VNĐ")
                    item_value.setFlags(item_value.flags() & ~Qt.ItemIsEditable)
                    item_value.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    table.setItem(row, 3, item_value)

                except Exception as e:
                    print(f"Lỗi khi hiển thị top sản phẩm hàng {row}: {str(e)}")
                    continue

            table.resizeColumnsToContents()
            
        except Exception as e:
            print(f"Lỗi khi hiển thị top sản phẩm tồn kho: {str(e)}")

    def search_inventory(self):
        """Tìm kiếm và lọc báo cáo tồn kho"""
        try:
            # Lấy các bộ lọc từ UI
            category = self.view.cboCategory.currentText() if hasattr(self.view, 'cboCategory') else None
            brand = self.view.cboBrand.currentText() if hasattr(self.view, 'cboBrand') else None
            stock_status = self.view.cboStockStatus.currentText() if hasattr(self.view, 'cboStockStatus') else None

            # Tải dữ liệu với các bộ lọc
            self.load_inventory_data(category, brand, stock_status)

            QMessageBox.information(self.view.parent_widget, "Thành công", 
                                    f"Đã tải báo cáo với {len(self.current_data)} sản phẩm.")

        except Exception as e:
            print(f"Lỗi khi tìm kiếm: {str(e)}")
            QMessageBox.warning(self.view.parent_widget, "Lỗi", f"Không thể tìm kiếm: {str(e)}")

    def reset_filters(self):
        """Đặt lại tất cả bộ lọc"""
        try:
            if hasattr(self.view, 'cboCategory'):
                self.view.cboCategory.setCurrentIndex(0)
            if hasattr(self.view, 'cboBrand'):
                self.view.cboBrand.setCurrentIndex(0)
            if hasattr(self.view, 'cboStockStatus'):
                self.view.cboStockStatus.setCurrentIndex(0)
            
            # Tải lại dữ liệu
            self.load_inventory_data()
            
            QMessageBox.information(self.view.parent_widget, "Thành công", "Đã đặt lại bộ lọc.")
            
        except Exception as e:
            print(f"Lỗi khi đặt lại bộ lọc: {str(e)}")
            QMessageBox.warning(self.view.parent_widget, "Lỗi", f"Không thể đặt lại bộ lọc: {str(e)}")

    def export_to_excel(self):
        """Xuất báo cáo tồn kho ra file Excel"""
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            filename, _ = QFileDialog.getSaveFileName(
                self.view.parent_widget, "Xuất báo cáo", f"BaoCaoTonKho_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel Files (*.xlsx);;All Files (*)"
            )

            if filename:
                # Tạo workbook
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Báo cáo tồn kho"

                # Thiết lập header
                headers = ["Mã SP", "Tên sản phẩm", "Danh mục", "Thương hiệu", 
                          "Giá nhập", "Giá bán", "Tồn kho", "Giá trị tồn", "Tỷ lệ %"]
                ws.append(["BÁO CÁO TỒN KHO", "", "", "", "", "", "", "", ""])
                ws.merge_cells('A1:I1')
                ws['A1'].font = Font(bold=True, size=14)
                ws['A1'].alignment = Alignment(horizontal='center')

                # Thêm ngày tạo báo cáo
                ws.append(["Ngày tạo: " + datetime.now().strftime('%d/%m/%Y %H:%M:%S'), "", "", "", "", "", "", "", ""])
                ws.merge_cells('A2:I2')

                # Header hàng
                ws.append([])
                ws.append(headers)

                # Style header
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=5, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # Thêm dữ liệu
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                # Tính tổng giá trị trước
                total_inventory_value = sum(
                    float(p.selling_price or 0) * int(p.quantity or 0) for p in self.current_data
                ) if self.current_data else 1

                for row_idx, product in enumerate(self.current_data, start=6):
                    selling_price = float(product.selling_price or 0)
                    quantity = int(product.quantity or 0)
                    purchase_price = float(product.purchase_price or 0)
                    inventory_value = selling_price * quantity
                    percentage = (inventory_value / total_inventory_value * 100) if total_inventory_value > 0 else 0

                    row_data = [
                        str(product.product_id or ""),
                        str(product.name or ""),
                        str(product.category or ""),
                        str(product.brand or ""),
                        purchase_price,
                        selling_price,
                        quantity,
                        inventory_value,
                        percentage
                    ]

                    ws.append(row_data)

                    # Định dạng cell
                    for col in range(1, len(headers) + 1):
                        cell = ws.cell(row=row_idx, column=col)
                        cell.border = border
                        cell.alignment = Alignment(horizontal='right' if col > 4 else 'left')

                        # Định dạng số
                        if col in [5, 6, 8]:  # Các cột giá
                            cell.number_format = '#,##0'
                        elif col == 9:  # Cột tỷ lệ
                            cell.number_format = '0.00"%"'

                # Thêm tổng cộng
                total_row = len(self.current_data) + 6
                ws.append([])
                total_purchase_value = sum(
                    float(p.purchase_price or 0) * int(p.quantity or 0) for p in self.current_data
                )
                total_quantity = sum(int(p.quantity or 0) for p in self.current_data)
                total_selling_value = sum(
                    float(p.selling_price or 0) * int(p.quantity or 0) for p in self.current_data
                )
                ws.append(["TỔNG CỘNG", "", "", "",
                          total_purchase_value,
                          "",
                          total_quantity,
                          total_selling_value,
                          "100%"])

                # Style tổng cộng
                total_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                total_font = Font(bold=True)
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=total_row + 2, column=col)
                    cell.fill = total_fill
                    cell.font = total_font
                    cell.border = border

                # Điều chỉnh độ rộng cột
                column_widths = [12, 30, 15, 15, 15, 15, 12, 18, 12]
                for idx, width in enumerate(column_widths, 1):
                    ws.column_dimensions[get_column_letter(idx)].width = width

                # Lưu file
                wb.save(filename)
                QMessageBox.information(self.view.parent_widget, "Thành công", 
                                        f"Đã xuất báo cáo thành công!\nFile: {filename}")

        except ImportError:
            # Nếu không có openpyxl, xuất CSV thay thế
            self.export_to_csv()
        except Exception as e:
            QMessageBox.warning(self.view.parent_widget, "Lỗi", f"Không thể xuất báo cáo: {str(e)}")

    def export_to_csv(self):
        """Xuất báo cáo tồn kho ra file CSV"""
        try:
            filename, _ = QFileDialog.getSaveFileName(
                self.view.parent_widget, "Xuất báo cáo", f"BaoCaoTonKho_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                "CSV Files (*.csv);;All Files (*)"
            )

            if filename:
                with open(filename, 'w', newline='', encoding='utf-8-sig') as file:
                    writer = csv.writer(file)
                    
                    # Header
                    writer.writerow(["BÁO CÁO TỒN KHO"])
                    writer.writerow(["Ngày tạo: " + datetime.now().strftime('%d/%m/%Y %H:%M:%S')])
                    writer.writerow([])
                    
                    # Column headers
                    headers = ["Mã SP", "Tên sản phẩm", "Danh mục", "Thương hiệu",
                              "Giá nhập", "Giá bán", "Tồn kho", "Giá trị tồn", "Tỷ lệ %"]
                    writer.writerow(headers)

                    # Data
                    total_inventory_value = sum(
                        float(p.selling_price or 0) * int(p.quantity or 0) for p in self.current_data
                    ) if self.current_data else 1

                    for product in self.current_data:
                        selling_price = float(product.selling_price or 0)
                        purchase_price = float(product.purchase_price or 0)
                        quantity = int(product.quantity or 0)
                        inventory_value = selling_price * quantity
                        percentage = (inventory_value / total_inventory_value * 100) if total_inventory_value > 0 else 0

                        writer.writerow([
                            str(product.product_id or ""),
                            str(product.name or ""),
                            str(product.category or ""),
                            str(product.brand or ""),
                            f"{purchase_price:,.0f}",
                            f"{selling_price:,.0f}",
                            quantity,
                            f"{inventory_value:,.0f}",
                            f"{percentage:.2f}%"
                        ])

                    # Summary
                    writer.writerow([])
                    total_purchase_value = sum(
                        float(p.purchase_price or 0) * int(p.quantity or 0) for p in self.current_data
                    )
                    total_quantity = sum(int(p.quantity or 0) for p in self.current_data)
                    total_selling_value = sum(
                        float(p.selling_price or 0) * int(p.quantity or 0) for p in self.current_data
                    )
                    writer.writerow(["TỔNG CỘNG",
                                    "",
                                    "",
                                    "",
                                    f"{total_purchase_value:,.0f}",
                                    "",
                                    total_quantity,
                                    f"{total_selling_value:,.0f}",
                                    "100%"])

                QMessageBox.information(self.view.parent_widget, "Thành công",
                                        f"Đã xuất báo cáo thành công!\nFile: {filename}")

        except Exception as e:
            QMessageBox.warning(self.view.parent_widget, "Lỗi", f"Không thể xuất báo cáo: {str(e)}")

    def print_report(self):
        """In báo cáo tồn kho"""
        try:
            from PyQt5.QtPrintSupport import QPrinter, QPrintDialog
            from PyQt5.QtCore import Qt
            from PyQt5.QtGui import QTextDocument
            import io

            # Tạo PDF document
            printer = QPrinter(QPrinter.HighResolution)
            dialog = QPrintDialog(printer, self.view.parent_widget)

            if dialog.exec_() == QFileDialog.Accepted:
                # Tạo HTML content
                html_content = self.generate_html_report()

                # Tạo document và in
                document = QTextDocument()
                document.setHtml(html_content)
                document.print(printer)

                QMessageBox.information(self.view.parent_widget, "Thành công", "Báo cáo đã được gửi tới máy in!")

        except Exception as e:
            QMessageBox.warning(self.view.parent_widget, "Lỗi", f"Không thể in báo cáo: {str(e)}")

    def generate_html_report(self):
        """Tạo nội dung HTML cho báo cáo"""
        html = f"""
        <html>
        <head>
            <meta charset="utf-8">
            <title>Báo cáo tồn kho</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                h1 {{ text-align: center; color: #333; }}
                .report-date {{ text-align: center; color: #666; margin-bottom: 20px; }}
                table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
                th {{ background-color: #366092; color: white; padding: 12px; text-align: left; border: 1px solid #999; }}
                td {{ padding: 10px; border: 1px solid #ddd; }}
                tr:nth-child(even) {{ background-color: #f9f9f9; }}
                .total-row {{ background-color: #FFFF99; font-weight: bold; }}
                .number {{ text-align: right; }}
                .center {{ text-align: center; }}
            </style>
        </head>
        <body>
            <h1>BÁO CÁO TỒN KHO</h1>
            <p class="report-date">Ngày tạo: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}</p>
            
            <table>
                <tr>
                    <th>Mã SP</th>
                    <th>Tên sản phẩm</th>
                    <th>Danh mục</th>
                    <th>Thương hiệu</th>
                    <th>Giá nhập</th>
                    <th>Giá bán</th>
                    <th>Tồn kho</th>
                    <th>Giá trị tồn</th>
                    <th>Tỷ lệ %</th>
                </tr>
        """

        total_inventory_value = sum(
            float(p.selling_price or 0) * int(p.quantity or 0) for p in self.current_data
        ) if self.current_data else 1

        for product in self.current_data:
            selling_price = float(product.selling_price or 0)
            purchase_price = float(product.purchase_price or 0)
            quantity = int(product.quantity or 0)
            inventory_value = selling_price * quantity
            percentage = (inventory_value / total_inventory_value * 100) if total_inventory_value > 0 else 0

            html += f"""
                <tr>
                    <td>{str(product.product_id or "")}</td>
                    <td>{str(product.name or "")}</td>
                    <td>{str(product.category or "")}</td>
                    <td>{str(product.brand or "")}</td>
                    <td class="number">{purchase_price:,.0f} VNĐ</td>
                    <td class="number">{selling_price:,.0f} VNĐ</td>
                    <td class="center">{quantity}</td>
                    <td class="number">{inventory_value:,.0f} VNĐ</td>
                    <td class="center">{percentage:.2f}%</td>
                </tr>
            """

        # Tổng cộng
        total_qty = sum(int(p.quantity or 0) for p in self.current_data)
        total_val = sum(float(p.selling_price or 0) * int(p.quantity or 0) for p in self.current_data)

        html += f"""
                <tr class="total-row">
                    <td colspan="6"><strong>TỔNG CỘNG</strong></td>
                    <td class="center"><strong>{total_qty}</strong></td>
                    <td class="number"><strong>{total_val:,.0f} VNĐ</strong></td>
                    <td class="center"><strong>100%</strong></td>
                </tr>
            </table>
        </body>
        </html>
        """

        return html
