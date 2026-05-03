from PyQt5.QtWidgets import QMessageBox, QTableWidgetItem, QAbstractItemView, QVBoxLayout, QLabel, QTextEdit, QDialog, QDialogButtonBox
from PyQt5.QtCore import Qt
from models.order import Order
from models.warranty import Warranty


class OrderController:
    """Controller xử lý logic quản lý đơn hàng"""

    def __init__(self, view):
        self.view = view
        self.setup_connections()
        self.load_orders()

    def setup_connections(self):
        if hasattr(self.view, "tableOrders"):
            self.view.tableOrders.itemSelectionChanged.connect(self.on_order_selected)
            self.view.tableOrders.setSelectionBehavior(QAbstractItemView.SelectRows)
            self.view.tableOrders.setSelectionMode(QAbstractItemView.SingleSelection)
            self.view.tableOrders.setEditTriggers(QAbstractItemView.NoEditTriggers)
            self.view.tableOrders.setAlternatingRowColors(True)
            self.view.tableOrders.setShowGrid(True)

        if hasattr(self.view, "btnViewDetail"):
            self.view.btnViewDetail.clicked.connect(self.view_order_details)
        
        if hasattr(self.view, "btnCancel"):
            self.view.btnCancel.clicked.connect(self.cancel_order)
        
        if hasattr(self.view, "btnPrint"):
            self.view.btnPrint.clicked.connect(self.print_order)
        
        if hasattr(self.view, "btnSearch"):
            self.view.btnSearch.clicked.connect(self.search_orders)
        
        if hasattr(self.view, "btnReset"):
            self.view.btnReset.clicked.connect(self.reset_filters)
        
        if hasattr(self.view, "btnExport"):
            self.view.btnExport.clicked.connect(self.export_orders)

    def load_orders(self):
        try:
            orders = Order.get_all()
            self.display_orders(orders)
        except Exception as e:
            QMessageBox.warning(None, "Lỗi", f"Không thể tải danh sách hóa đơn: {str(e)}")

    def display_orders(self, orders):
        if not hasattr(self.view, "tableOrders"):
            return

        table = self.view.tableOrders
        table.clear()
        table.setRowCount(len(orders))
        table.setColumnCount(9)
        table.setHorizontalHeaderLabels([
            "Mã HD", "Ngày tạo", "Khách hàng", "SĐT", "Tổng tiền", "Đã thanh toán", "Còn nợ", "Trạng thái", "Thao tác"
        ])

        for row, order in enumerate(orders):
            table.setItem(row, 0, QTableWidgetItem(order.order_number))
            table.setItem(row, 1, QTableWidgetItem(order.created_at or ""))
            table.setItem(row, 2, QTableWidgetItem(order.customer_name or ""))
            table.setItem(row, 3, QTableWidgetItem(order.customer_phone or ""))
            table.setItem(row, 4, QTableWidgetItem(self.format_currency(order.total_amount)))
            table.setItem(row, 5, QTableWidgetItem(self.format_currency(order.paid_amount)))
            table.setItem(row, 6, QTableWidgetItem(self.format_currency(max(order.total_amount - order.paid_amount, 0))))
            table.setItem(row, 7, QTableWidgetItem(order.status or ""))
            table.setItem(row, 8, QTableWidgetItem("👁️ Xem"))

        table.resizeColumnsToContents()

    def on_order_selected(self):
        if not hasattr(self.view, "tableOrders") or not hasattr(self.view, "tableOrderDetails"):
            return

        selected_rows = self.view.tableOrders.selectionModel().selectedRows()
        if not selected_rows:
            self.display_order_details([])
            return

        row = selected_rows[0].row()
        order_number = self.view.tableOrders.item(row, 0).text()
        if not order_number:
            self.display_order_details([])
            return

        order = Order.get_by_id(order_number)
        if not order:
            self.display_order_details([])
            return

        self.display_order_details(order.items)

    def display_order_details(self, items):
        table = self.view.tableOrderDetails
        table.clear()
        table.setRowCount(len(items))
        table.setColumnCount(5)
        table.setHorizontalHeaderLabels([
            "Mã SP", "Tên sản phẩm", "Số lượng", "Đơn giá", "Thành tiền"
        ])

        for row, item in enumerate(items):
            table.setItem(row, 0, QTableWidgetItem(item["product_id"]))
            table.setItem(row, 1, QTableWidgetItem(item["product_name"]))
            table.setItem(row, 2, QTableWidgetItem(str(item["quantity"])))
            table.setItem(row, 3, QTableWidgetItem(self.format_currency(item["unit_price"])))
            table.setItem(row, 4, QTableWidgetItem(self.format_currency(item["total_price"])))

        table.resizeColumnsToContents()

    def cancel_order(self):
        if not hasattr(self.view, "tableOrders"):
            return

        selected_rows = self.view.tableOrders.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.warning(None, "Cảnh báo", "Vui lòng chọn một hóa đơn để hủy.")
            return

        row = selected_rows[0].row()
        order_number = self.view.tableOrders.item(row, 0).text()
        order_status = self.view.tableOrders.item(row, 7).text()

        if order_status == "Đã hủy":
            QMessageBox.warning(None, "Cảnh báo", "Hóa đơn này đã hủy rồi.")
            return

        reply = QMessageBox.question(
            None,
            "Xác nhận hủy hóa đơn",
            f"Bạn có chắc chắn muốn hủy hóa đơn {order_number}?",
            QMessageBox.Yes | QMessageBox.No,
        )
        
        if reply != QMessageBox.Yes:
            return

        try:
            order = Order.get_by_id(order_number)
            if not order:
                QMessageBox.warning(None, "Lỗi", "Không tìm thấy hóa đơn.")
                return

            order.update_status("Đã hủy")
            self.load_orders()
            QMessageBox.information(None, "Thành công", f"Hủy hóa đơn {order_number} thành công!")
        except Exception as e:
            QMessageBox.warning(None, "Lỗi", f"Hủy hóa đơn thất bại: {str(e)}")

    def print_order(self):
        if not hasattr(self.view, "tableOrders"):
            return

        selected_rows = self.view.tableOrders.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.warning(None, "Cảnh báo", "Vui lòng chọn một hóa đơn để in.")
            return

        row = selected_rows[0].row()
        order_number = self.view.tableOrders.item(row, 0).text()

        try:
            order = Order.get_by_id(order_number)
            if not order:
                QMessageBox.warning(None, "Lỗi", "Không tìm thấy hóa đơn.")
                return

            self.show_print_dialog(order)
        except Exception as e:
            QMessageBox.warning(None, "Lỗi", f"Không thể tải dữ liệu hóa đơn: {str(e)}")

    def show_print_dialog(self, order):
        from PyQt5.QtPrintSupport import QPrinter, QPrintDialog
        from PyQt5.QtGui import QTextDocument
        
        # Tạo tài liệu HTML cho hóa đơn
        html_content = f"""
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
                th, td {{ border: 1px solid #ddd; padding: 10px; text-align: left; }}
                th {{ background-color: #f0f0f0; font-weight: bold; }}
                .header {{ text-align: center; margin-bottom: 20px; }}
                .info {{ margin: 10px 0; }}
                .total {{ font-weight: bold; font-size: 16px; text-align: right; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h2>ĐơN HÓA ĐơN MUA HÀNG</h2>
            </div>
            <div class="info">
                <p><b>Mã hóa đơn:</b> {order.order_number}</p>
                <p><b>Ngày tạo:</b> {order.created_at}</p>
                <p><b>Tên khách hàng:</b> {order.customer_name}</p>
                <p><b>SĐT:</b> {order.customer_phone}</p>
                <p><b>Trạng thái:</b> {order.status}</p>
            </div>
            <table>
                <tr>
                    <th>Mã SP</th>
                    <th>Tên sản phẩm</th>
                    <th>Số lượng</th>
                    <th>Đơn giá</th>
                    <th>Thành tiền</th>
                </tr>
        """
        
        for item in order.items:
            html_content += f"""
                <tr>
                    <td>{item['product_id']}</td>
                    <td>{item['product_name']}</td>
                    <td>{item['quantity']}</td>
                    <td>{self.format_currency(item['unit_price'])}</td>
                    <td>{self.format_currency(item['total_price'])}</td>
                </tr>
            """
        
        html_content += f"""
            </table>
            <div class="total" style="margin-top: 20px;">
                <p>Tổng cộng: {self.format_currency(order.total_amount)}</p>
                <p>Đã thanh toán: {self.format_currency(order.paid_amount)}</p>
                <p>Còn nợ: {self.format_currency(max(order.total_amount - order.paid_amount, 0))}</p>
            </div>
        </body>
        </html>
        """
        
        # Tạo printer
        printer = QPrinter()
        dialog = QPrintDialog(printer)
        
        if dialog.exec() == QMessageBox.Accepted:
            doc = QTextDocument()
            doc.setHtml(html_content)
            doc.print(printer)
            QMessageBox.information(None, "Thành công", "In hóa đơn thành công!")

    def view_order_details(self):
        if not hasattr(self.view, "tableOrders"):
            return

        selected_rows = self.view.tableOrders.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.warning(None, "Cảnh báo", "Vui lòng chọn một hóa đơn để xem chi tiết.")
            return

        row = selected_rows[0].row()
        order_number = self.view.tableOrders.item(row, 0).text()

        try:
            order = Order.get_by_id(order_number)
            if not order:
                QMessageBox.warning(None, "Lỗi", "Không tìm thấy hóa đơn.")
                return

            self.show_order_details_dialog(order)
        except Exception as e:
            QMessageBox.warning(None, "Lỗi", f"Không thể tải chi tiết hóa đơn: {str(e)}")

    def show_order_details_dialog(self, order):
        dialog = QDialog()
        dialog.setWindowTitle(f"Chi tiết hóa đơn {order.order_number}")
        dialog.resize(600, 400)

        layout = QVBoxLayout()

        # Thông tin cơ bản
        info_label = QLabel(f"""
        <b>Mã hóa đơn:</b> {order.order_number}<br>
        <b>Tên khách hàng:</b> {order.customer_name}<br>
        <b>ID nhân viên:</b> {order.employee_id}<br>
        <b>Ngày đặt:</b> {order.order_date}<br>
        <b>Tổng tiền:</b> {self.format_currency(order.total_amount)}<br>
        <b>Trạng thái:</b> {order.status}
        """)
        info_label.setWordWrap(True)
        layout.addWidget(info_label)

        # Chi tiết sản phẩm
        details_label = QLabel("<b>Chi tiết sản phẩm:</b>")
        layout.addWidget(details_label)

        details_text = QTextEdit()
        details_text.setReadOnly(True)

        details = "STT\tMã SP\tTên sản phẩm\t\tSL\tĐơn giá\t\tThành tiền\n"
        details += "-" * 80 + "\n"

        for i, item in enumerate(order.items, 1):
            details += f"{i}\t{item['product_id']}\t{item['product_name'][:20]}\t\t{item['quantity']}\t{self.format_currency(item['unit_price'])}\t\t{self.format_currency(item['total_price'])}\n"

        details_text.setPlainText(details)
        layout.addWidget(details_text)

        # Buttons
        buttons = QDialogButtonBox(QDialogButtonBox.Ok)
        buttons.accepted.connect(dialog.accept)
        layout.addWidget(buttons)

        dialog.setLayout(layout)
        dialog.exec()

    def search_orders(self):
        try:
            order_code = ""
            customer_name = ""
            status = ""
            date_from = ""
            date_to = ""
            
            if hasattr(self.view, "txtOrderCode"):
                order_code = self.view.txtOrderCode.text().strip()
            
            if hasattr(self.view, "txtCustomer"):
                customer_name = self.view.txtCustomer.text().strip()
            
            if hasattr(self.view, "cboStatus"):
                status_text = self.view.cboStatus.currentText()
                if status_text and status_text != "Tất cả":
                    status = status_text
            
            if hasattr(self.view, "dateFrom"):
                date_from = self.view.dateFrom.date().toString("yyyy-MM-dd")
            
            if hasattr(self.view, "dateTo"):
                date_to = self.view.dateTo.date().toString("yyyy-MM-dd")
            
            orders = Order.search_by_filters(order_code, customer_name, status, date_from, date_to)
            self.display_orders(orders)
        except Exception as e:
            QMessageBox.warning(None, "Lỗi", f"Tìm kiếm thất bại: {str(e)}")

    def reset_filters(self):
        if hasattr(self.view, "txtOrderCode"):
            self.view.txtOrderCode.clear()
        
        if hasattr(self.view, "txtCustomer"):
            self.view.txtCustomer.clear()
        
        if hasattr(self.view, "cboStatus"):
            self.view.cboStatus.setCurrentIndex(0)
        
        if hasattr(self.view, "dateFrom"):
            from PyQt5.QtCore import QDate
            self.view.dateFrom.setDate(QDate(2000, 1, 1))
        
        if hasattr(self.view, "dateTo"):
            from PyQt5.QtCore import QDate
            self.view.dateTo.setDate(QDate.currentDate())
        
        self.load_orders()

    def export_orders(self):
        try:
            import csv
            from PyQt5.QtWidgets import QFileDialog
            
            file_path, _ = QFileDialog.getSaveFileName(
                None,
                "Lưu hóa đơn",
                "",
                "CSV Files (*.csv);;Excel Files (*.xlsx)"
            )
            
            if not file_path:
                return
            
            table = self.view.tableOrders
            rows = table.rowCount()
            cols = table.columnCount()
            
            if file_path.endswith('.csv'):
                with open(file_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
                    writer = csv.writer(csvfile)
                    
                    headers = []
                    for col in range(cols):
                        item = table.horizontalHeaderItem(col)
                        headers.append(item.text() if item else "")
                    writer.writerow(headers)
                    
                    for row in range(rows):
                        row_data = []
                        for col in range(cols):
                            item = table.item(row, col)
                            row_data.append(item.text() if item else "")
                        writer.writerow(row_data)
                
                QMessageBox.information(None, "Thành công", f"Xuất dữ liệu thành công: {file_path}")
            
            elif file_path.endswith('.xlsx'):
                try:
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, PatternFill
                    
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Hóa đơn"
                    
                    headers = []
                    for col in range(cols):
                        item = table.horizontalHeaderItem(col)
                        headers.append(item.text() if item else "")
                    
                    ws.append(headers)
                    
                    header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF")
                    
                    for col_num, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col_num)
                        cell.fill = header_fill
                        cell.font = header_font
                    
                    for row in range(rows):
                        row_data = []
                        for col in range(cols):
                            item = table.item(row, col)
                            row_data.append(item.text() if item else "")
                        ws.append(row_data)
                    
                    wb.save(file_path)
                    QMessageBox.information(None, "Thành công", f"Xuất dữ liệu thành công: {file_path}")
                except ImportError:
                    QMessageBox.warning(None, "Lỗi", "Cần cài đặt openpyxl để xuất XLSX.\nHay chọn CSV thay thế.")
        except Exception as e:
            QMessageBox.warning(None, "Lỗi", f"Xuất dữ liệu thất bại: {str(e)}")

    @staticmethod
    def format_currency(value):
        return f"{value:,.0f} VNĐ"
