from PyQt5.QtWidgets import (QMessageBox, QTableWidgetItem, QLineEdit, QPushButton, 
                           QComboBox, QDateEdit, QFileDialog)
from PyQt5.QtCore import Qt, QDate
from datetime import datetime
from models.salary import Salary
from models.employee import Employee
from models.database import Database


class SalaryController:
    """Controller xử lý logic tính lương"""
    
    def __init__(self, view):
        self.view = view
        self.current_salary = None
        self.initialize_ui()
        self.setup_connections()
        self.load_salaries()
    
    def initialize_ui(self):
        """Khởi tạo các giá trị UI"""
        # Set current month and year
        now = datetime.now()
        if hasattr(self.view, "spinMonth"):
            self.view.spinMonth.setValue(now.month)
        if hasattr(self.view, "spinYear"):
            self.view.spinYear.setValue(now.year)
        
        # Populate employee combobox
        if hasattr(self.view, "cboEmployee"):
            self.view.cboEmployee.clear()
            employees = Employee.get_all()
            self.view.cboEmployee.addItem("Chọn nhân viên", None)
            for emp in employees:
                self.view.cboEmployee.addItem(f"{emp.employee_id} - {emp.full_name}", emp.employee_id)
        
        # Populate status combobox
        if hasattr(self.view, "cboStatus"):
            self.view.cboStatus.clear()
            self.view.cboStatus.addItems(["Nháp", "Phê duyệt", "Đã thanh toán"])

    def setup_connections(self):
        """Thiết lập kết nối các signal"""
        if hasattr(self.view, "btnCalculate"):
            self.view.btnCalculate.clicked.connect(self.calculate_salary)
        if hasattr(self.view, "btnApprove"):
            self.view.btnApprove.clicked.connect(self.approve_salary)
        if hasattr(self.view, "btnPay"):
            self.view.btnPay.clicked.connect(self.pay_salary)
        if hasattr(self.view, "btnRefresh"):
            self.view.btnRefresh.clicked.connect(self.load_salaries)
        if hasattr(self.view, "tableSalaries"):
            self.view.tableSalaries.itemSelectionChanged.connect(self.on_salary_selected)

    def load_salaries(self):
        """Tải danh sách bảng lương"""
        try:
            month = self.view.spinMonth.value() if hasattr(self.view, "spinMonth") else datetime.now().month
            year = self.view.spinYear.value() if hasattr(self.view, "spinYear") else datetime.now().year
            
            salaries = Salary.get_by_month(month, year) if month and year else Salary.get_all()
            self.display_salaries(salaries)
        except Exception as e:
            QMessageBox.warning(None, "Lỗi", f"Không thể tải danh sách lương: {str(e)}")

    def display_salaries(self, salaries):
        """Hiển thị danh sách bảng lương lên bảng"""
        if not hasattr(self.view, "tableSalaries"):
            return

        table = self.view.tableSalaries
        table.setRowCount(len(salaries))
        table.setColumnCount(9)
        table.setHorizontalHeaderLabels([
            "Mã", "Nhân viên", "Tháng", "Năm", "Lương cơ bản",
            "Lương bruto", "Lương ròng", "Trạng thái", "Ghi chú"
        ])

        for row, salary in enumerate(salaries):
            emp = Employee.get_by_id(salary.employee_id)
            emp_name = f"{emp.full_name}" if emp else salary.employee_id
            
            table.setItem(row, 0, QTableWidgetItem(salary.salary_id))
            table.setItem(row, 1, QTableWidgetItem(emp_name))
            table.setItem(row, 2, QTableWidgetItem(str(salary.month)))
            table.setItem(row, 3, QTableWidgetItem(str(salary.year)))
            table.setItem(row, 4, QTableWidgetItem(f"{salary.base_salary:,.0f}"))
            table.setItem(row, 5, QTableWidgetItem(f"{salary.gross_salary:,.0f}"))
            table.setItem(row, 6, QTableWidgetItem(f"{salary.net_salary:,.0f}"))
            table.setItem(row, 7, QTableWidgetItem(salary.status))
            table.setItem(row, 8, QTableWidgetItem(salary.note))

        table.resizeColumnsToContents()

    def calculate_salary(self):
        """Tính lương cho nhân viên"""
        try:
            if hasattr(self.view, "cboEmployee"):
                employee_id = self.view.cboEmployee.currentData()
                if not employee_id:
                    QMessageBox.warning(None, "Cảnh báo", "Vui lòng chọn nhân viên!")
                    return
            else:
                return

            month = self.view.spinMonth.value() if hasattr(self.view, "spinMonth") else datetime.now().month
            year = self.view.spinYear.value() if hasattr(self.view, "spinYear") else datetime.now().year
            
            # Kiểm tra bảng lương đã tồn tại
            existing = Salary.get_by_month(month, year)
            if any(s.employee_id == employee_id for s in existing):
                QMessageBox.warning(None, "Cảnh báo", "Bảng lương tháng này đã tồn tại!")
                return

            emp = Employee.get_by_id(employee_id)
            if not emp:
                QMessageBox.warning(None, "Lỗi", "Không tìm thấy nhân viên!")
                return

            overtime_hours = float(self.view.spinOvertimeHours.value()) if hasattr(self.view, "spinOvertimeHours") else 0
            bonus = float(self.view.spinBonus.value()) if hasattr(self.view, "spinBonus") else 0
            deduction = float(self.view.spinDeduction.value()) if hasattr(self.view, "spinDeduction") else 0

            # Tạo bảng lương
            salary = Salary.create(
                employee_id=employee_id,
                month=month,
                year=year,
                base_salary=emp.base_salary,
                overtime_hours=overtime_hours,
                bonus=bonus,
                deduction=deduction
            )

            QMessageBox.information(None, "Thành công", f"Đã tính lương thành công! Mã: {salary.salary_id}")
            self.load_salaries()

        except Exception as e:
            QMessageBox.warning(None, "Lỗi", f"Không thể tính lương: {str(e)}")

    def approve_salary(self):
        """Phê duyệt bảng lương"""
        if not self.current_salary:
            QMessageBox.warning(None, "Cảnh báo", "Vui lòng chọn bảng lương!")
            return

        if self.current_salary.approve():
            QMessageBox.information(None, "Thành công", "Đã phê duyệt bảng lương!")
            self.load_salaries()
        else:
            QMessageBox.warning(None, "Cảnh báo", "Chỉ phê duyệt được bảng lương ở trạng thái 'Nháp'!")

    def pay_salary(self):
        """Thanh toán bảng lương"""
        if not self.current_salary:
            QMessageBox.warning(None, "Cảnh báo", "Vui lòng chọn bảng lương!")
            return

        if self.current_salary.pay():
            QMessageBox.information(None, "Thành công", "Đã thanh toán bảng lương!")
            self.load_salaries()
        else:
            QMessageBox.warning(None, "Cảnh báo", "Chỉ thanh toán được bảng lương đã phê duyệt!")

    def on_salary_selected(self):
        """Xử lý khi chọn bảng lương"""
        if not hasattr(self.view, "tableSalaries"):
            return

        current_row = self.view.tableSalaries.currentRow()
        if current_row >= 0:
            salary_id = self.view.tableSalaries.item(current_row, 0).text()
            self.current_salary = Salary.get_by_id(salary_id)

    def export_excel(self):
        """Xuất bảng lương ra Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill

            month = self.view.spinMonth.value() if hasattr(self.view, "spinMonth") else datetime.now().month
            year = self.view.spinYear.value() if hasattr(self.view, "spinYear") else datetime.now().year
            
            salaries = Salary.get_by_month(month, year)
            if not salaries:
                QMessageBox.warning(None, "Cảnh báo", "Không có bảng lương để xuất!")
                return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Lương {month}/{year}"

            # Header
            headers = ["Mã", "Nhân viên", "Lương cơ bản", "Tăng ca", "Lương tăng ca",
                      "Thưởng", "Khấu trừ", "Lương bruto", "Lương ròng", "Trạng thái"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")

            # Data
            for row_idx, salary in enumerate(salaries, 2):
                emp = Employee.get_by_id(salary.employee_id)
                ws.cell(row=row_idx, column=1).value = salary.salary_id
                ws.cell(row=row_idx, column=2).value = emp.full_name if emp else salary.employee_id
                ws.cell(row=row_idx, column=3).value = salary.base_salary
                ws.cell(row=row_idx, column=4).value = salary.overtime_hours
                ws.cell(row=row_idx, column=5).value = salary.overtime_salary
                ws.cell(row=row_idx, column=6).value = salary.bonus
                ws.cell(row=row_idx, column=7).value = salary.deduction
                ws.cell(row=row_idx, column=8).value = salary.gross_salary
                ws.cell(row=row_idx, column=9).value = salary.net_salary
                ws.cell(row=row_idx, column=10).value = salary.status

            for column in ws.columns:
                ws.column_dimensions[column[0].column_letter].width = 15

            filename = f"luong_{month}_{year}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(filename)
            QMessageBox.information(None, "Thành công", f"Đã xuất file: {filename}")

        except ImportError:
            QMessageBox.warning(None, "Lỗi", "Vui lòng cài đặt: pip install openpyxl")
        except Exception as e:
            QMessageBox.warning(None, "Lỗi", f"Không thể xuất file: {str(e)}")